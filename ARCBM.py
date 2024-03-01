# haha this was alot of work Sir.
# Made by Don Michael Tranquada of ODCS1.

# All of my library imports down below.
import sys
import os
from PyQt5.QtWidgets import QWidget, QApplication, QVBoxLayout, QLabel, QLineEdit, QTextEdit, QPushButton, QFileDialog, QMessageBox, QDesktopWidget
from PyQt5.QtGui import QIcon, QFont, QFontDatabase, QPixmap
from PyQt5.QtCore import Qt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from shutil import copyfile
icon_path = 'arcbm.ico'
font_path = 'C:/Users/PHILLIP JACOBUS/OneDrive/Desktop/bioformproj/fonts/MinecraftRegular-Bmg3.otf'

class ARCBM(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.check_database_exists()
        self.image_path = None

    def init_ui(self):
        self.setWindowTitle('A Really Crappy Bio Maker by Don Michael Tranquada')
        self.setWindowState(Qt.WindowFullScreen)
        self.setGeometry(QDesktopWidget().screenGeometry())
        self.setStyleSheet("background-color: #1e1e1e; color: white;")

        font_id = QFontDatabase.addApplicationFont(font_path)
        if font_id != -1:
            font_family = QFontDatabase.applicationFontFamilies(font_id)[0]
            font = QFont(font_family, 16)
            QApplication.setFont(font)

        self.setWindowIcon(QIcon(icon_path))

        layout = QVBoxLayout()

        personal_bio_label = QLabel('A Really Crappy Bio Maker by Don Michael Tranquada', self)
        personal_bio_label.setStyleSheet("font-size: 20px;")
        personal_bio_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(personal_bio_label)

        self.first_name_label = QLabel('First Name:', self)
        self.first_name_label.setStyleSheet("font-size: 16px;")
        self.first_name_entry = QLineEdit(self)
        self.first_name_entry.setStyleSheet("font-size: 16px; background-color: #2b2b2b; color: white; border: none;")
        layout.addWidget(self.first_name_label)
        layout.addWidget(self.first_name_entry)

        self.last_name_label = QLabel('Last Name:', self)
        self.last_name_label.setStyleSheet("font-size: 16px;")
        self.last_name_entry = QLineEdit(self)
        self.last_name_entry.setStyleSheet("font-size: 16px; background-color: #2b2b2b; color: white; border: none;")
        layout.addWidget(self.last_name_label)
        layout.addWidget(self.last_name_entry)

        self.dob_label = QLabel('Date of Birth:', self)
        self.dob_label.setStyleSheet("font-size: 16px;")
        self.dob_entry = QLineEdit(self)
        self.dob_entry.setStyleSheet("font-size: 16px; background-color: #2b2b2b; color: white; border: none;")
        layout.addWidget(self.dob_label)
        layout.addWidget(self.dob_entry)

        self.email_label = QLabel('Email:', self)
        self.email_label.setStyleSheet("font-size: 16px;")
        self.email_entry = QLineEdit(self)
        self.email_entry.setStyleSheet("font-size: 16px; background-color: #2b2b2b; color: white; border: none;")
        layout.addWidget(self.email_label)
        layout.addWidget(self.email_entry)

        self.number_label = QLabel('Phone Number:', self)
        self.number_label.setStyleSheet("font-size: 16px;")
        self.number_entry = QLineEdit(self)
        self.number_entry.setStyleSheet("font-size: 16px; background-color: #2b2b2b; color: white; border: none;")
        layout.addWidget(self.number_label)
        layout.addWidget(self.number_entry)

        education_label = QLabel('Education', self)
        education_label.setStyleSheet("font-size: 20px;")
        layout.addWidget(education_label)

        self.education_text = QTextEdit(self)
        self.education_text.setStyleSheet("font-size: 16px; background-color: #2b2b2b; color: white; border: none;")
        layout.addWidget(self.education_text)

        work_exp_label = QLabel('Work Experience', self)
        work_exp_label.setStyleSheet("font-size: 20px;")
        layout.addWidget(work_exp_label)

        self.work_exp_text = QTextEdit(self)
        self.work_exp_text.setStyleSheet("font-size: 16px; background-color: #2b2b2b; color: white; border: none;")
        layout.addWidget(self.work_exp_text)

        self.upload_button = QPushButton('Upload Image', self)
        self.upload_button.setStyleSheet("background-color: #444; color: white; border: none;")
        self.upload_button.clicked.connect(self.upload_image)
        layout.addWidget(self.upload_button)

        self.image_label = QLabel(self)
        layout.addWidget(self.image_label)

        self.save_button = QPushButton('Save Data', self)
        self.save_button.setStyleSheet("background-color: #444; color: white; border: none;")
        self.save_button.clicked.connect(self.save_data)
        layout.addWidget(self.save_button)

        self.setLayout(layout)
        
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_F and event.modifiers() == Qt.ControlModifier:
            if self.isFullScreen():
                self.showNormal()
            else:
                self.showFullScreen()
        elif event.key() == Qt.Key_S and event.modifiers() == Qt.ControlModifier:
            self.save_data()
        elif event.key() == Qt.Key_Q and event.modifiers() == Qt.ControlModifier:
            self.close()

    def check_database_exists(self):
        self.database_path = "A Really Crappy Bio Maker Database/arcbmlist.xlsx"
        self.image_dir = "A Really Crappy Bio Maker Database/arcbmpics"

        if not os.path.exists(self.image_dir):
            os.makedirs(self.image_dir)

        if not os.path.exists(self.database_path):
            new_workbook = Workbook()
            new_workbook.save(self.database_path)
            QMessageBox.information(self, 'Database Created',
                                    f'The ARCBM database excel file has been created in the directory: '
                                    f'{os.path.dirname(os.path.abspath(self.database_path))}.')
        else:
            QMessageBox.information(self, 'Database Exists',
                                    f'The ARCBM database excel file already exists in the directory: '
                                    f'{os.path.dirname(os.path.abspath(self.database_path))}.')

    def upload_image(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filename, _ = QFileDialog.getOpenFileName(self, "Upload Image", "", "Image Files (*.jpg *.jpeg *.png)", options=options)
        if filename:
            self.image_path = filename
            pixmap = QPixmap(self.image_path)
            scaled_pixmap = pixmap.scaled(300, 300, Qt.KeepAspectRatio)
            self.image_label.setAlignment(Qt.AlignCenter)
            self.image_label.setPixmap(scaled_pixmap)

    def save_data(self):
        first_name = self.first_name_entry.text()
        last_name = self.last_name_entry.text()
        age = self.dob_entry.text()
        email = self.email_entry.text()
        number = self.number_entry.text()
        education = self.education_text.toPlainText()
        work_experience = self.work_exp_text.toPlainText()

        save_msg_box = QMessageBox()
        save_msg_box.setWindowTitle('Save Data')
        save_msg_box.setText('Do you want to create a new personal file or add to the ARCBM database?')
        save_msg_box.addButton('New', QMessageBox.AcceptRole)
        save_msg_box.addButton('Append', QMessageBox.AcceptRole)
        save_msg_box.addButton('Cancel', QMessageBox.RejectRole)
        save_msg_box.exec()

        if save_msg_box.clickedButton().text() == 'Cancel':
            return

        if save_msg_box.clickedButton().text() == 'Append':
            self.save_to_database(first_name, last_name, age, email, number, education, work_experience)
        elif save_msg_box.clickedButton().text() == 'New':
            self.create_personal_file(first_name, last_name, age, email, number, education, work_experience)

    def save_to_database(self, first_name, last_name, age, email, number, education, work_experience):
        workbook = load_workbook(self.database_path)
        worksheet = workbook.active

        bold_italic_font = Font(bold=True, italic=True)
        if worksheet.max_row == 1:  # If the first row is empty add the headers
            worksheet.append(['First Name', 'Last Name', 'Age', 'Email', 'Phone Number', 'Education', 'Work Experience'])
            for cell in worksheet[1]:
                cell.font = bold_italic_font

        worksheet.append([first_name, last_name, age, email, number, education, work_experience])
        workbook.save(self.database_path)

        if self.image_path:
            new_image_name = f"{first_name}_{last_name}.png"
            new_image_path = os.path.join(self.image_dir, new_image_name)
            copyfile(self.image_path, new_image_path)

        QMessageBox.information(self, 'Success', 'Data saved successfully to ARCBM database.')

    def create_personal_file(self, first_name, last_name, age, email, number, education, work_experience):
        personal_file_path = os.path.join(os.path.dirname(self.database_path), f"{first_name}_{last_name}.xlsx")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['First Name', 'Last Name', 'Age', 'Email', 'Phone Number', 'Education', 'Work Experience'])
        worksheet.append([first_name, last_name, age, email, number, education, work_experience])
        bold_italic_font = Font(bold=True, italic=True)
        for cell in worksheet[1]:
            cell.font = bold_italic_font
        workbook.save(personal_file_path)

        if self.image_path:
            new_image_name = f"{first_name}_{last_name}.png"
            new_image_path = os.path.join(self.image_dir, new_image_name)
            copyfile(self.image_path, new_image_path)

        QMessageBox.information(self, 'Success', f'Personal file {first_name}_{last_name}.xlsx created successfully.')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ARCBM()
    window.show()
    sys.exit(app.exec_())